# 爬虫模块
import os
from urllib import parse

import bs4
import requests
from openpyxl import Workbook, load_workbook
from xpinyin import Pinyin


# 下载网页,生成html静态文件
def pageDown(url, fileName):
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0"
    }
    try:
        response = requests.get(url, headers=headers)
        response.encoding = 'gbk'
        with open('./output/' + fileName, mode='w', encoding='gbk') as file:
            file.write(response.text)
            pass
        pass
    except:
        print('下载失败')
        raise
    pass


# 解析网页,并保存到excel文件
def pageParse(fileName, cityName, jobName, saveFileName):
    try:
        with open('./output/' + fileName, mode='r', encoding='gbk') as file:
            content = file.read()
            pass
        soup = bs4.BeautifulSoup(content, 'lxml')
        jobList = soup.select('#resultList .el')
        new_jobList = []
        for item in jobList:
            jobName = item.select('.t1')[0].get_text(strip=True)
            company = item.select(".t2")[0].get_text(strip=True)
            city = item.select(".t3")[0].get_text(strip=True)
            money = item.select(".t4")[0].get_text(strip=True)
            time = item.select(".t5")[0].get_text(strip=True)
            row = {
                "jobName": jobName,
                "company": company,
                "city": city,
                "money": money,
                "time": time
            }
            new_jobList.append(row)
            pass
        saveExcel(saveFileName, cityName, jobName, new_jobList)
        pass
    except:
        print('解析失败')
        raise
    pass


# 保存到excel文件
def saveExcel(saveFileName, cityName, jobName, jobList):
    titleList = ['职位名', '公司名', '工作地点', '薪资', '发布时间']
    try:
        if os.path.isfile('./output/' + saveFileName):
            wb = load_workbook('./output/' + saveFileName)
            sheetList = wb.get_sheet_names()
            if cityName not in sheetList:
                new_sheet = wb.create_sheet(cityName)  # 新增sheet
                new_sheet.append(titleList)
                for item1 in jobList[1:]:  # 新增数据
                    row = [item1['jobName'], item1['company'], item1['city'], item1['money'], item1['time']]
                    new_sheet.append(row)
                    wb.save('./output/' + saveFileName)
                    pass
                pass
            else:
                ws = wb[cityName]
                for item in jobList[1:]:  # 新增数据
                    row = [item['jobName'], item['company'], item['city'], item['money'], item['time']]
                    ws.append(row)
                    wb.save('./output/' + saveFileName)
                    pass
                pass
            pass
        else:
            book = Workbook()  # 初始化excel对象
            sheet = book.create_sheet(cityName)  # 新增sheet
            sheet.append(titleList)
            for item in jobList[1:]:  # 新增数据
                row = [item['jobName'], item['company'], item['city'], item['money'], item['time']]
                sheet.append(row)
                pass
            book.save('./output/' + saveFileName)
            pass
        pass
    except:
        print('保存失败')
        raise
    pass


# 分页、分类爬取取51job职位数据，并解析
def getBypage(jobName, cityName, pageSize):
    p = Pinyin()
    new_jobName = parse.quote(parse.quote(jobName))
    fileName = jobName
    cityId = ''
    for item in cityList:
        if item['cityName'] == cityName:
            cityId = item['id']
            pass
        pass
    try:
        for i in range(pageSize):
            pageDown(
                'https://search.51job.com/list/0' + str(cityId) + '00,000000,0000,00,9,99,' + new_jobName + ',2,' + str(
                    i + 1) + '.html?lang=c&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&ord_field=0&dibiaoid=0&line=&welfare=',
                'html/' + fileName + 'Data.html')
            pageParse('html/' + fileName + 'Data.html', cityName, jobName, 'excel/' + fileName + '.xlsx')
            pass
        print('分页读取成功')
        pass
    except:
        print('分页读取失败，城市名或页数错误')
        raise
    pass


if __name__ == '__main__':
    cityList = [{'id': 100, 'cityName': '北京'}, {'id': 200, 'cityName': '上海'},
                {'id': 302, 'cityName': '广州'}, {'id': 400, 'cityName': '深圳'},
                {'id': 902, 'cityName': '成都'}]
    for item in cityList:
        print('城市:' + item['cityName'])
        pass
    jobName = input('请输入要爬取的职位名称:')
    cityName = input('请输入地点:')
    pageSize = input('请输入页数:')
    getBypage(jobName, cityName, int(pageSize))
    pass
